import openpyxl
wb = openpyxl.load_workbook('/Users/mac/Desktop/宁氏项目清单.xlsx')
for name in wb.sheetnames:
    print(f"\n===== Sheet: {name} =====")
    ws = wb[name]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        print([str(c) if c is not None else '' for c in row])
