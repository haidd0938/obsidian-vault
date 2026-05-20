import openpyxl
wb = openpyxl.load_workbook('/Users/mac/Desktop/宁氏项目清单.xlsx', data_only=True)
ws = wb.active
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    vals = [v if v is not None else '' for v in row]
    print(vals)
